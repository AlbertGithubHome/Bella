#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Date     : 2020-3-15 08:55:51
# @Author   : Albert Shi
# @Link     : http://blog.csdn.net/albertsh
# @Github   : https://github.com/AlbertGithubHome
__author__ = 'AlbertS'
# @Subject  : 使用 openpyxl 库操作 Excel

# 注意：1. 这种方式最好用来操作 .xlsx 格式的文件
#       2. .xls 格式文件无法打开，但是可以保存，新建保存后会提示格式错误，但可以看到基础数据
#       3. 处理不了带有宏的表格，强行存储为 .xlsx 格式化后宏会全部丢失

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill

FILE_PATH_ROOT = '../data/'
FILE_PATH_A = 'sampleA.xlsx'
FILE_PATH_B = 'sampleB.xls'
FILE_PATH_C = 'sampleC.xlsx'
FILE_PATH_D = 'sampleD.xls'
FILE_PATH_M = 'sampleM.xlsm'

def read_update_excel(file_name):
    # 加载Excel表
    wb = load_workbook(FILE_PATH_ROOT + file_name)
    # 打印sheet数量
    print('sheet count:', len(wb.sheetnames))
    # 打印所有sheet名字
    print('sheet name list:', wb.sheetnames)
    # 获取第一个sheet对象
    ws = wb[wb.sheetnames[0]]
    # 打印sheet表行数和列数
    print('rows count:', ws.max_row, 'cols count:', ws.max_column)
    # 更新单元格A1的内容
    ws['A1'] = 'this is A1'
    # 在第二行位置插入一行
    ws.insert_rows(2)
    # 删除第五行
    ws.delete_rows(5)
    # 获取单元格对象，对应B2单元格
    cell = ws.cell(2,2)
    # 设置单元格内容
    cell.value = 'this is B2'
    # 修改字体格式为粗体
    cell.font = Font(bold=True)
    # 修改单元格格式
    cell.fill = PatternFill("solid", fgColor="F0CDCD")

    # 保存原文件或另存一个文件
    wb.save(file_name)

def write_new_excel(file_name):
    # 创建一个excel文档
    wb = Workbook()
    # 获得当前激活的sheet对象
    ws = wb.active
    # 给A2单元格赋值
    ws['A2'] = 'This is A2 cell'
    # 一行添加多列数据
    ws.append([1, 2, 'hello'])
    # 添加新的sheet
    ws = wb.create_sheet(title='NewInfo',index=0)
    # 设置单元格的值
    ws['A1'] = 'This is new sheet'

    # 保存excel
    wb.save(file_name)

if __name__ == '__main__':
    read_update_excel(FILE_PATH_A)
    #read_update_excel(FILE_PATH_B)
    read_update_excel(FILE_PATH_M)
    write_new_excel(FILE_PATH_C)
    write_new_excel(FILE_PATH_D)