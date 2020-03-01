# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill


def w():
    #创建一个工作薄对象,也就是创建一个excel文档
    wb = Workbook()

    #指定当前显示（活动）的sheet对象
    ws = wb.active

    # 给A1单元格赋值
    ws['A1'] = 42

    # 一行添加多列数据
    ws.append([1, 2, 3])

    # 保存excel
    wb.save("sample.xlsx")

def rw():
    wb = load_workbook('origin_excel.xlsx')

    print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]

    print(ws.max_row)
    print(ws.max_column)

    #ws['A1'] = 'aaaa'
    ws.insert_rows(2)

    # b22 = ws.cell(2,2)
    # b22.value = 'zzz'
    # b22.fill = PatternFill("solid", fgColor="00CDCD")
    for x in range(ws.max_column):
        cell = ws.cell(2,x+1)
        cell.value = 'zzz'
        cell.fill = PatternFill("solid", fgColor="00CDCD")

    ws.insert_rows(3)
    for x in range(ws.max_column):
        cell = ws.cell(3,x+1)
        cell.value = '测试'
        cell.fill = PatternFill("solid", fgColor="AABBCC")


    # 保存excel
    wb.save("sample.xlsx")

if __name__ == '__main__':
    rw()